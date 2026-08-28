from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Annotated, Any
from urllib.parse import quote

from fastapi import Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MAX_IMPORT_BYTES = 8 * 1024 * 1024
MAX_IMPORT_ROWS = 5000

_HEADER_ALIASES = {
    "person_name": {"الاسم الثلاثي", "الاسم", "اسم المستفيد", "person_name", "person name", "name"},
    "amount": {"المبلغ", "مبلغ السلفة", "amount", "advance amount"},
    "advance_month": {"الشهر", "شهر السلفة", "advance_month", "advance month", "month"},
    "notes": {"ملاحظات", "الملاحظات", "notes", "note"},
}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().replace("_", " ").lower()
    return re.sub(r"\s+", " ", text)


def _header_map(values: list[Any]) -> dict[str, int]:
    normalized = [_normalize_header(value) for value in values]
    result: dict[str, int] = {}
    for key, aliases in _HEADER_ALIASES.items():
        normalized_aliases = {_normalize_header(alias) for alias in aliases}
        for index, value in enumerate(normalized):
            if value in normalized_aliases:
                result[key] = index
                break
    return result


def _normalize_month(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m")
    text = str(value or "").strip()
    if not text:
        raise ValueError("الشهر مطلوب")
    text = text.replace("/", "-").replace(".", "-")
    match = re.fullmatch(r"(\d{4})-(\d{1,2})(?:-\d{1,2})?", text)
    if not match:
        raise ValueError("صيغة الشهر يجب أن تكون YYYY-MM")
    year = int(match.group(1))
    month = int(match.group(2))
    if year < 1900 or year > 2200 or month < 1 or month > 12:
        raise ValueError("قيمة الشهر غير صحيحة")
    return f"{year:04d}-{month:02d}"


def _normalize_amount_minor(value: Any) -> int:
    arabic = "٠١٢٣٤٥٦٧٨٩"
    persian = "۰۱۲۳۴۵۶۷۸۹"
    text = str(value if value is not None else "").strip()
    text = "".join(str(arabic.index(ch)) if ch in arabic else ch for ch in text)
    text = "".join(str(persian.index(ch)) if ch in persian else ch for ch in text)
    text = text.replace("٬", "").replace(",", "").replace(" ", "").replace("٫", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    try:
        amount = Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        raise ValueError("المبلغ غير صحيح")
    minor = int(amount * 100)
    if minor <= 0:
        raise ValueError("المبلغ يجب أن يكون أكبر من صفر")
    return minor


def _safe_text(value: Any, max_length: int) -> str:
    text = "" if value is None else str(value).strip()
    if len(text) > max_length:
        raise ValueError(f"القيمة أطول من الحد المسموح ({max_length})")
    return text


def _query_advances(core: Any, conn: Any, *, q: str = "", status: str | None = None, month: str | None = None):
    where: list[str] = []
    params: list[Any] = []
    if q.strip():
        where.append("(a.person_name LIKE ? OR a.notes LIKE ? OR a.advance_month LIKE ?)")
        term = f"%{q.strip()}%"
        params.extend([term, term, term])
    if status == "active":
        where.append("a.remaining_amount_minor > 0")
    elif status == "paid":
        where.append("a.remaining_amount_minor = 0")
    if month:
        normalized_month = _normalize_month(month)
        where.append("a.advance_month = ?")
        params.append(normalized_month)
    sql = core.ADVANCE_SELECT
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY a.advance_month DESC, a.updated_at DESC, a.id DESC"
    return conn.execute(sql, params).fetchall()


def _style_header(ws, row_number: int = 1) -> None:
    fill = PatternFill("solid", fgColor="0E6B4F")
    font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(bottom=Side(style="thin", color="CFC7B2"))
    for cell in ws[row_number]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[row_number].height = 24


def _fit_columns(ws, *, min_width: int = 11, max_width: int = 38) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(column_index)].width = max(min_width, min(max_width, max_len + 3))


def _summary_sheet(core: Any, workbook: Workbook, rows: list[Any], filters: dict[str, Any]) -> None:
    ws = workbook.active
    ws.title = "ملخص"
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "تقرير السلف — ملخص"
    ws["A1"].font = Font(size=18, bold=True, color="0E6B4F")
    ws.merge_cells("A1:D1")
    ws["A3"] = "تاريخ التصدير"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"] = "بحث"
    ws["B4"] = filters.get("q") or "الكل"
    ws["A5"] = "الحالة"
    ws["B5"] = {"active": "قائمة", "paid": "مسددة بالكامل"}.get(filters.get("status"), "الكل")
    ws["A6"] = "الشهر"
    ws["B6"] = filters.get("month") or "كل الأشهر"

    total_amount = sum(int(row["amount_minor"]) for row in rows)
    total_remaining = sum(int(row["remaining_amount_minor"]) for row in rows)
    total_paid = total_amount - total_remaining
    active_count = sum(1 for row in rows if int(row["remaining_amount_minor"]) > 0)
    paid_count = len(rows) - active_count
    metrics = [
        ("عدد السلف", len(rows)),
        ("السلف القائمة", active_count),
        ("المسددة بالكامل", paid_count),
        ("إجمالي مبالغ السلف", float(Decimal(total_amount) / Decimal(100))),
        ("إجمالي المسدد", float(Decimal(total_paid) / Decimal(100))),
        ("إجمالي المتبقي", float(Decimal(total_remaining) / Decimal(100))),
    ]
    start = 9
    for offset, (label, value) in enumerate(metrics):
        row_num = start + offset
        ws.cell(row_num, 1, label)
        ws.cell(row_num, 2, value)
        ws.cell(row_num, 1).font = Font(bold=True)
        if offset >= 3:
            ws.cell(row_num, 2).number_format = '#,##0.00'
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 24


def _data_sheet(core: Any, workbook: Workbook, conn: Any, rows: list[Any]) -> None:
    ws = workbook.create_sheet("السلف")
    ws.sheet_view.rightToLeft = True
    headers = [
        "رقم السلفة", "الاسم الثلاثي", "الشهر", "مبلغ السلفة", "المسدد", "المتبقي",
        "الحالة", "عدد التسديدات", "الملاحظات", "أنشأها", "تاريخ الإنشاء", "آخر تعديل",
    ]
    ws.append(headers)
    _style_header(ws)
    for row in rows:
        item = core._advance_dict(conn, row)
        ws.append([
            item["id"], item["person_name"], item["advance_month"], float(item["amount"]),
            float(item["paid_amount"]), float(item["remaining_amount"]),
            "مسددة بالكامل" if item["status"] == "paid" else "قائمة",
            item["payment_count"], item["notes"] or "", item["created_by_name"] or "",
            item["created_at"], item["updated_at"],
        ])
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = '#,##0.00'
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _fit_columns(ws)
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["I"].width = 36


def _template_sheet(workbook: Workbook) -> None:
    ws = workbook.create_sheet("نموذج الاستيراد")
    ws.sheet_view.rightToLeft = True
    ws.append(["الاسم الثلاثي", "المبلغ", "الشهر", "ملاحظات"])
    _style_header(ws)
    ws["A2"] = ""
    ws["B2"] = ""
    ws["C2"] = ""
    ws["D2"] = ""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:D1"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 42
    ws["A4"] = "تعليمات"
    ws["A4"].font = Font(bold=True, color="0E6B4F")
    ws["A5"] = "الشهر بصيغة YYYY-MM مثل 2026-08. كل صف ينشئ سلفة جديدة دون إنشاء تسديدات تاريخية."
    ws.merge_cells("A5:D5")
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[5].height = 34


def _workbook_bytes(core: Any, conn: Any, rows: list[Any], filters: dict[str, Any]) -> bytes:
    workbook = Workbook()
    _summary_sheet(core, workbook, rows, filters)
    _data_sheet(core, workbook, conn, rows)
    _template_sheet(workbook)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _template_bytes() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _template_sheet(workbook)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _xlsx_response(payload: bytes, filename: str) -> StreamingResponse:
    encoded = quote(filename)
    return StreamingResponse(
        io.BytesIO(payload),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


def register_advance_excel_routes(core: Any) -> None:
    if getattr(core.app.state, "advance_excel_routes_registered", False):
        return
    core.app.state.advance_excel_routes_registered = True

    @core.app.get("/api/advance-excel/export")
    def export_advances_excel(
        user: Annotated[core.CurrentUser, Depends(core.require("documents.read"))],
        q: str = Query(default="", max_length=160),
        status: str | None = Query(default=None, pattern=r"^(active|paid)$"),
        month: str | None = Query(default=None, max_length=10),
    ):
        conn = core.connect()
        try:
            core._advance_page_access(conn, user)
            rows = _query_advances(core, conn, q=q, status=status, month=month)
            payload = _workbook_bytes(core, conn, rows, {"q": q, "status": status, "month": month})
            return _xlsx_response(payload, f"advances-{datetime.now().strftime('%Y-%m-%d')}.xlsx")
        finally:
            conn.close()

    @core.app.get("/api/advance-excel/template")
    def advance_excel_template(
        user: Annotated[core.CurrentUser, Depends(core.require("documents.read"))],
    ):
        conn = core.connect()
        try:
            core._advance_page_access(conn, user)
        finally:
            conn.close()
        return _xlsx_response(_template_bytes(), "advances-import-template.xlsx")

    @core.app.get("/api/advance-excel/report/{advance_id}")
    def export_advance_report_excel(
        advance_id: int,
        user: Annotated[core.CurrentUser, Depends(core.require("documents.read"))],
    ):
        conn = core.connect()
        try:
            core._advance_page_access(conn, user)
            row = conn.execute(core.ADVANCE_SELECT + " WHERE a.id=?", (advance_id,)).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="السلفة غير موجودة")
            payload = _workbook_bytes(core, conn, [row], {"q": "", "status": None, "month": row["advance_month"]})
            return _xlsx_response(payload, f"advance-{advance_id}-{row['advance_month']}.xlsx")
        finally:
            conn.close()

    @core.app.post("/api/advance-excel/import")
    async def import_advances_excel(
        request: Request,
        user: Annotated[core.CurrentUser, Depends(core.require("documents.create"))],
    ):
        raw = await request.body()
        if not raw:
            raise HTTPException(status_code=422, detail="اختر ملف Excel أولاً")
        if len(raw) > MAX_IMPORT_BYTES:
            raise HTTPException(status_code=413, detail="حجم ملف Excel أكبر من الحد المسموح 8 MB")
        try:
            workbook = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        except Exception as exc:
            raise HTTPException(status_code=422, detail="تعذر قراءة الملف. استخدم ملف Excel بصيغة .xlsx") from exc

        sheet = workbook["نموذج الاستيراد"] if "نموذج الاستيراد" in workbook.sheetnames else workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_values = list(next(rows_iter))
        except StopIteration:
            raise HTTPException(status_code=422, detail="ملف Excel فارغ")
        columns = _header_map(header_values)
        required = ["person_name", "amount", "advance_month"]
        missing = [key for key in required if key not in columns]
        if missing:
            labels = {"person_name": "الاسم الثلاثي", "amount": "المبلغ", "advance_month": "الشهر"}
            raise HTTPException(status_code=422, detail="الأعمدة المطلوبة غير موجودة: " + "، ".join(labels[key] for key in missing))

        conn = core.connect()
        imported = 0
        skipped = 0
        errors: list[dict[str, Any]] = []
        try:
            core._advance_page_access(conn, user)
            with core.transaction(conn, immediate=True):
                for excel_row_number, values in enumerate(rows_iter, start=2):
                    if excel_row_number > MAX_IMPORT_ROWS + 1:
                        errors.append({"row": excel_row_number, "error": f"تم إيقاف القراءة عند {MAX_IMPORT_ROWS} صف"})
                        break
                    values = list(values)
                    if not any(value not in (None, "") for value in values):
                        continue
                    try:
                        person_name = _safe_text(values[columns["person_name"]] if columns["person_name"] < len(values) else "", 160)
                        if len(person_name) < 3:
                            raise ValueError("الاسم الثلاثي مطلوب ويجب أن يكون 3 أحرف على الأقل")
                        amount_minor = _normalize_amount_minor(values[columns["amount"]] if columns["amount"] < len(values) else None)
                        advance_month = _normalize_month(values[columns["advance_month"]] if columns["advance_month"] < len(values) else None)
                        notes = ""
                        if "notes" in columns and columns["notes"] < len(values):
                            notes = _safe_text(values[columns["notes"]], 2000)
                        duplicate = conn.execute(
                            "SELECT id FROM advances WHERE TRIM(person_name)=? AND amount_minor=? AND advance_month=? LIMIT 1",
                            (person_name, amount_minor, advance_month),
                        ).fetchone()
                        if duplicate:
                            skipped += 1
                            continue
                        now = core.utc_iso()
                        cursor = conn.execute(
                            """
                            INSERT INTO advances(person_name, amount_minor, notes, advance_month, remaining_amount_minor,
                                                 created_by, updated_by, created_at, updated_at)
                            VALUES(?,?,?,?,?,?,?,?,?)
                            """,
                            (person_name, amount_minor, notes, advance_month, amount_minor, user.id, user.id, now, now),
                        )
                        core.audit(
                            conn,
                            user_id=user.id,
                            action="advance.create",
                            entity_type="advance",
                            entity_id=int(cursor.lastrowid),
                            details={
                                "person_name": person_name,
                                "amount": core._minor_to_money(amount_minor),
                                "month": advance_month,
                                "source": "excel_import",
                                "excel_row": excel_row_number,
                            },
                        )
                        imported += 1
                    except Exception as exc:
                        errors.append({"row": excel_row_number, "error": str(exc)[:240]})
                        if len(errors) >= 100:
                            break
            return {
                "ok": True,
                "imported": imported,
                "skipped": skipped,
                "errors": errors,
                "error_count": len(errors),
            }
        finally:
            conn.close()
